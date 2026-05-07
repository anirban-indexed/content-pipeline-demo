"""
context_loader.py — Reads all context files from the client profile and returns
their content as text to be passed to Claude with every session.

Handles: .docx (python-docx), .xlsx (openpyxl), .csv (pandas),
         .txt and .md (plain text)
"""

from __future__ import annotations
import os
import pandas as pd
import openpyxl
from docx import Document


def read_docx(path: str) -> str:
    """Extract all text from a .docx file."""
    try:
        doc = Document(path)
        return "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
    except Exception as e:
        return f"[ERROR reading {os.path.basename(path)}: {e}]"


def read_xlsx(path: str) -> str:
    """Extract all sheets from a .xlsx file as readable text."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        output = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output.append(f"--- Sheet: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    output.append(row_text)
        return "\n".join(output)
    except Exception as e:
        return f"[ERROR reading {os.path.basename(path)}: {e}]"


def read_csv(path: str) -> str:
    """Read a CSV file and return it as readable text."""
    try:
        df = pd.read_csv(path)
        return df.to_string(index=False)
    except Exception as e:
        return f"[ERROR reading {os.path.basename(path)}: {e}]"


def read_text(path: str) -> str:
    """Read a plain text or markdown file."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        return f"[ERROR reading {os.path.basename(path)}: {e}]"


def _read_file(path: str) -> str:
    """Dispatch to the correct reader based on file extension."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".docx":
        return read_docx(path)
    elif ext == ".xlsx":
        return read_xlsx(path)
    elif ext == ".csv":
        return read_csv(path)
    elif ext in (".txt", ".md"):
        return read_text(path)
    else:
        return f"[Unsupported file type: {os.path.basename(path)}]"


def extract_internal_link_pool(path: str, domain: str) -> str:
    """
    Extract URLs and topics from the content optimization CSV
    and return as a clean list for internal linking recommendations.
    Filters rows whose URL contains `domain`.
    """
    try:
        df = pd.read_csv(path, header=None, skiprows=2)
        output = []
        for _, row in df.iterrows():
            url = str(row.iloc[0]).strip()
            topic = str(row.iloc[1]).strip()
            if (url.startswith("https://")
                    and domain in url
                    and topic not in ("nan", "", "Topic")):
                output.append(f"- {topic} → {url}")
        return "\n".join(output) if output else "No internal link data available."
    except Exception as e:
        return f"[ERROR extracting internal link pool: {e}]"


def load_all_context(config, profile: dict) -> dict:
    """
    Load all context files specified in the client profile and return as a
    dict of label -> text content.

    Args:
        config:  The config module (used for ROOT_DIR; not for file paths).
        profile: The loaded client profile dict (from config.load_client_profile).
    """
    context: dict = {}
    ctx_files = profile.get("_context_files", {})
    domain = profile.get("domain", "")

    # Editorial handbook
    handbook_path = ctx_files.get("editorial_handbook", "")
    if handbook_path and handbook_path != "TODO" and os.path.exists(handbook_path):
        context["editorial_handbook"] = _read_file(handbook_path)
    else:
        context["editorial_handbook"] = ""

    # State page info (optional — Veriheal-specific; skipped if not in profile)
    state_page_path = ctx_files.get("state_page_info", "")
    if state_page_path and state_page_path != "TODO" and os.path.exists(state_page_path):
        context["state_page_info"] = _read_file(state_page_path)
    else:
        context["state_page_info"] = ""

    # Example articles
    context["example_articles"] = {}
    for path in ctx_files.get("example_articles", []):
        if path and path != "TODO" and os.path.exists(path):
            label = os.path.splitext(os.path.basename(path))[0]
            context["example_articles"][label] = _read_file(path)

    # Internal link pool — derived from CSV if present; txt file otherwise
    csv_path = ctx_files.get("content_optimization_csv", "")
    if csv_path and csv_path != "TODO" and os.path.exists(csv_path):
        context["internal_link_pool"] = extract_internal_link_pool(csv_path, domain)
    else:
        txt_path = os.path.join(
            profile.get("_client_dir", ""),
            "context",
            "internal_links.txt",
        )
        if os.path.exists(txt_path):
            raw = read_text(txt_path)
            lines = [l for l in raw.splitlines() if l.strip() and not l.strip().startswith("#")]
            context["internal_link_pool"] = "\n".join(lines) if lines else "No internal link data available."
        else:
            context["internal_link_pool"] = "No internal link data available."

    return context


def format_context_for_prompt(context: dict, profile: dict | None = None) -> str:
    """
    Format all loaded context into a single string block
    to be injected into the Claude prompt as reference material.
    """
    client_name = (profile or {}).get("client_name", "Client")
    sections = []

    if context.get("editorial_handbook"):
        sections.append(f"=== {client_name.upper()} EDITORIAL HANDBOOK ===")
        sections.append(context["editorial_handbook"])

    if context.get("state_page_info"):
        sections.append("\n=== STATE PAGE INFO ===")
        sections.append(context["state_page_info"])

    if context.get("example_articles"):
        sections.append("\n=== EXAMPLE ARTICLES (VOICE AND TONE REFERENCE) ===")
        for label, text in context["example_articles"].items():
            sections.append(f"\n--- {label} ---")
            sections.append(text)

    sections.append(f"\n=== {client_name.upper()} INTERNAL LINK POOL ===")
    sections.append("Only recommend URLs from this list for internal links in the brief.")
    sections.append("Do not recommend any URL not on this list. If no relevant URL exists, omit the internal link rather than inventing one.")
    sections.append(context.get("internal_link_pool", ""))

    return "\n".join(sections)
